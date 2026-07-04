import io
import json
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from app import db
from app.models import Teacher, Group, Student, Lesson, Task, Submission
from app.test_files import add_test
from werkzeug.security import generate_password_hash


@pytest.fixture
def teacher(app):
    with app.app_context():
        t = Teacher(username='admin', password_hash=generate_password_hash('secret123'))
        db.session.add(t)
        db.session.commit()
        return t


@pytest.fixture
def logged_in_client(client, teacher):
    client.post('/admin', data={'username': 'admin', 'password': 'secret123'})
    return client


@pytest.fixture
def full_setup(app):
    """Create group with 2 students, 1 lesson, 1 task, 2 tests, submissions."""
    with app.app_context():
        g = Group(name='7A_1gr')
        db.session.add(g)
        db.session.commit()

        s1 = Student(group_id=g.id, login='7A_1gr_1',
                      password_hash=generate_password_hash('a1'),
                      password_plain='a1', last_name='Иванов',
                      first_name='Иван', seq_number=1)
        s2 = Student(group_id=g.id, login='7A_1gr_2',
                      password_hash=generate_password_hash('a2'),
                      password_plain='a2', last_name='Петров',
                      first_name='Пётр', seq_number=2)
        db.session.add_all([s1, s2])

        l = Lesson(order_number=1, title='Урок 1')
        db.session.add(l)
        db.session.commit()

        t = Task(lesson_id=l.id, letter_index='A',
                 short_title='Сумма', problem_text='p',
                 input_description='i', output_description='o',
                 time_limit=2, memory_limit=256)
        db.session.add(t)
        db.session.commit()
        add_test(t.id, 1, '1 2\n', '3\n')
        add_test(t.id, 2, '3 4\n', '7\n')

        sub1 = Submission(student_id=s1.id, task_id=t.id,
                          code='print(1)', language='python',
                          test_results=json.dumps({
                              'status': 'ok', 'total_tests': 2,
                              'passed': 2, 'results': [],
                              'compile_error': None
                          }))
        sub2 = Submission(student_id=s2.id, task_id=t.id,
                          code='print(2)', language='python',
                          test_results=json.dumps({
                              'status': 'error', 'total_tests': 2,
                              'passed': 1, 'results': [],
                              'compile_error': None
                          }))
        db.session.add_all([sub1, sub2])
        db.session.commit()

        return g.id


# ── Results page ───────────────────────────────────────────────────────


class TestResultsPage:
    def test_get_results_page(self, logged_in_client):
        resp = logged_in_client.get('/teacher/results')
        assert resp.status_code == 200

    def test_results_page_shows_groups(self, logged_in_client, full_setup):
        resp = logged_in_client.get('/teacher/results')
        assert b'7A_1gr' in resp.data

    def test_results_with_group_selected(self, logged_in_client, full_setup):
        resp = logged_in_client.get(f'/teacher/results?group={full_setup}')
        assert resp.status_code == 200
        text = resp.data.decode('utf-8')
        assert 'Иванов' in text
        assert 'Петров' in text

    def test_results_shows_ok_for_full_solution(self, logged_in_client,
                                                  full_setup):
        resp = logged_in_client.get(f'/teacher/results?group={full_setup}')
        text = resp.data.decode('utf-8')
        assert 'OK' in text

    def test_results_shows_partial_score(self, logged_in_client,
                                           full_setup):
        resp = logged_in_client.get(f'/teacher/results?group={full_setup}')
        text = resp.data.decode('utf-8')
        assert '1/2' in text or '1 / 2' in text

    def test_results_empty_cell_no_submission(self, app, logged_in_client,
                                                full_setup):
        with app.app_context():
            s3 = Student(group_id=full_setup, login='7A_1gr_3',
                          password_hash=generate_password_hash('a3'),
                          password_plain='a3', last_name='Сидоров',
                          first_name='Сид', seq_number=3)
            db.session.add(s3)
            db.session.commit()
        resp = logged_in_client.get(f'/teacher/results?group={full_setup}')
        text = resp.data.decode('utf-8')
        assert 'Сидоров' in text

    def test_results_shows_task_column_header(self, logged_in_client,
                                                full_setup):
        resp = logged_in_client.get(f'/teacher/results?group={full_setup}')
        text = resp.data.decode('utf-8')
        assert 'Сумма' in text
        assert 'A' in text


# ── Export to Excel ────────────────────────────────────────────────────


class TestExportExcel:
    def test_export_returns_xlsx(self, logged_in_client, full_setup):
        resp = logged_in_client.get(
            f'/teacher/results/export?group={full_setup}')
        assert resp.status_code == 200
        assert 'spreadsheet' in resp.content_type or \
               'octet-stream' in resp.content_type

    def test_export_file_is_valid_xlsx(self, logged_in_client, full_setup):
        resp = logged_in_client.get(
            f'/teacher/results/export?group={full_setup}')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        assert ws.max_row >= 3
        assert ws.max_column >= 4

    def test_export_contains_student_names(self, logged_in_client,
                                             full_setup):
        resp = logged_in_client.get(
            f'/teacher/results/export?group={full_setup}')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        names = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    names.append(str(cell))
        full_text = ' '.join(names)
        assert 'Иванов' in full_text
        assert 'Петров' in full_text

    def test_export_contains_task_title(self, logged_in_client, full_setup):
        resp = logged_in_client.get(
            f'/teacher/results/export?group={full_setup}')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        header_cells = [str(c.value) for c in ws[1] if c.value]
        header_text = ' '.join(header_cells)
        assert 'Сумма' in header_text or 'A' in header_text

    def test_export_contains_ok_and_score(self, logged_in_client, full_setup):
        resp = logged_in_client.get(
            f'/teacher/results/export?group={full_setup}')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        all_cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_cells.append(str(cell))
        all_text = ' '.join(all_cells)
        assert 'OK' in all_text
        assert '1/2' in all_text

    def test_export_no_group_returns_400(self, logged_in_client):
        resp = logged_in_client.get('/teacher/results/export')
        assert resp.status_code == 400

    def test_export_invalid_group_returns_404(self, logged_in_client):
        resp = logged_in_client.get('/teacher/results/export?group=9999')
        assert resp.status_code == 404
