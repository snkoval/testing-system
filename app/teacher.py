import re
import io
import json

from flask import Blueprint, render_template, request, redirect, session, flash, abort, send_file

from app import db
from app.models import Teacher, Group, Student, Lesson, Task, Submission
from app.decorators import login_required_teacher
from app.utils import generate_password, generate_login, is_lesson_accessible
from app.test_files import get_tests, get_test_count, add_test, update_test, delete_test
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timezone

bp = Blueprint('teacher', __name__, url_prefix='/teacher')


@bp.route('/')
@login_required_teacher
def index():
    return render_template('teacher/dashboard.html')


@bp.route('/groups')
@login_required_teacher
def groups():
    groups = Group.query.order_by(Group.name).all()
    return render_template('teacher/groups.html', groups=groups)


@bp.route('/groups/create', methods=['GET', 'POST'])
@login_required_teacher
def create_group():
    if request.method == 'POST':
        class_number = request.form.get('class_number', '').strip()
        class_letter = request.form.get('class_letter', '').strip().upper()
        group_number = request.form.get('group_number', '').strip()

        if not re.match(r'^\d{1,2}$', class_number):
            flash('Номер класса должен быть 1–2 цифры', 'error')
            return render_template('teacher/create_group.html')
        if not re.match(r'^[А-ЯЁA-Z]$', class_letter):
            flash('Литера класса — одна заглавная буква', 'error')
            return render_template('teacher/create_group.html')
        if not re.match(r'^\d$', group_number):
            flash('Номер группы — одна цифра', 'error')
            return render_template('teacher/create_group.html')

        name = f'{class_number}{class_letter}_{group_number}gr'

        existing = Group.query.filter_by(name=name).first()
        if existing:
            flash(f'Группа {name} уже существует', 'error')
            return render_template('teacher/create_group.html')

        g = Group(name=name)
        db.session.add(g)
        db.session.commit()
        flash(f'Группа {name} создана', 'success')
        return redirect(f'/teacher/groups/{g.id}')

    return render_template('teacher/create_group.html')


@bp.route('/groups/<int:group_id>')
@login_required_teacher
def group_detail(group_id):
    group = db.session.get(Group, group_id)
    if group is None:
        abort(404)
    students = Student.query.filter_by(group_id=group_id).order_by(Student.seq_number).all()
    return render_template('teacher/group_detail.html', group=group, students=students)


@bp.route('/groups/<int:group_id>/add', methods=['POST'])
@login_required_teacher
def add_student(group_id):
    group = db.session.get(Group, group_id)
    if group is None:
        abort(404)

    last_name = request.form.get('last_name', '').strip()
    first_name = request.form.get('first_name', '').strip()

    if not last_name or not first_name:
        flash('Фамилия и имя обязательны', 'error')
        return redirect(f'/teacher/groups/{group_id}')

    max_seq = db.session.query(db.func.max(Student.seq_number)) \
        .filter_by(group_id=group_id).scalar() or 0
    seq_number = max_seq + 1
    login = generate_login(group.name, seq_number)
    plain = generate_password()

    s = Student(
        group_id=group_id,
        login=login,
        password_hash=generate_password_hash(plain),
        password_plain=plain,
        last_name=last_name,
        first_name=first_name,
        seq_number=seq_number
    )
    db.session.add(s)
    db.session.commit()
    flash(f'Ученик {login} добавлен. Пароль: {plain}', 'success')
    return redirect(f'/teacher/groups/{group_id}')


@bp.route('/groups/<int:group_id>/bulk-add', methods=['GET', 'POST'])
@login_required_teacher
def bulk_add(group_id):
    group = db.session.get(Group, group_id)
    if group is None:
        abort(404)

    if request.method == 'POST':
        students_data = request.form.get('students_data', '').strip()
        lines = [l.strip() for l in students_data.splitlines() if l.strip()]

        max_seq = db.session.query(db.func.max(Student.seq_number)) \
            .filter_by(group_id=group_id).scalar() or 0

        created = []
        for line in lines:
            parts = line.split(None, 1)
            if len(parts) < 2:
                continue
            last_name, first_name = parts[0], parts[1]
            max_seq += 1
            login = generate_login(group.name, max_seq)
            plain = generate_password()

            s = Student(
                group_id=group_id,
                login=login,
                password_hash=generate_password_hash(plain),
                password_plain=plain,
                last_name=last_name,
                first_name=first_name,
                seq_number=max_seq
            )
            db.session.add(s)
            created.append((login, plain, last_name, first_name))

        db.session.commit()
        return render_template('teacher/bulk_result.html', group=group, created=created)

    return render_template('teacher/bulk_add.html', group=group)


@bp.route('/groups/<int:group_id>/delete', methods=['POST'])
@login_required_teacher
def delete_group(group_id):
    group = db.session.get(Group, group_id)
    if group is None:
        abort(404)
    name = group.name
    db.session.delete(group)
    db.session.commit()
    flash(f'Группа {name} удалена', 'success')
    return redirect('/teacher/groups')


@bp.route('/lessons')
@login_required_teacher
def lessons():
    lessons = Lesson.query.order_by(Lesson.order_number).all()
    return render_template('teacher/lessons.html', lessons=lessons)


@bp.route('/lessons/create', methods=['GET', 'POST'])
@login_required_teacher
def create_lesson():
    if request.method == 'POST':
        order_number = request.form.get('order_number', '').strip()
        title = request.form.get('title', '').strip()
        theory_url = request.form.get('theory_url', '').strip()

        if not order_number or not order_number.isdigit():
            flash('Порядковый номер — целое число', 'error')
            return render_template('teacher/lesson_form.html')
        if not title:
            flash('Тема обязательна', 'error')
            return render_template('teacher/lesson_form.html')

        l = Lesson(
            order_number=int(order_number),
            title=title,
            theory_url=theory_url or None
        )
        db.session.add(l)
        db.session.commit()
        flash('Урок создан', 'success')
        return redirect('/teacher/lessons')

    return render_template('teacher/lesson_form.html')


@bp.route('/lessons/<int:lesson_id>/edit', methods=['GET', 'POST'])
@login_required_teacher
def edit_lesson(lesson_id):
    lesson = db.session.get(Lesson, lesson_id)
    if lesson is None:
        abort(404)

    if request.method == 'POST':
        order_number = request.form.get('order_number', '').strip()
        title = request.form.get('title', '').strip()
        theory_url = request.form.get('theory_url', '').strip()

        if not order_number or not order_number.isdigit():
            flash('Порядковый номер — целое число', 'error')
            return render_template('teacher/lesson_form.html', lesson=lesson)
        if not title:
            flash('Тема обязательна', 'error')
            return render_template('teacher/lesson_form.html', lesson=lesson)

        lesson.order_number = int(order_number)
        lesson.title = title
        lesson.theory_url = theory_url or None
        db.session.commit()
        flash('Урок обновлён', 'success')
        return redirect('/teacher/lessons')

    return render_template('teacher/lesson_form.html', lesson=lesson)


@bp.route('/lessons/<int:lesson_id>/access', methods=['POST'])
@login_required_teacher
def lesson_access(lesson_id):
    lesson = db.session.get(Lesson, lesson_id)
    if lesson is None:
        abort(404)

    action = request.form.get('action', '')

    if action == 'open':
        access_days_raw = request.form.get('access_days', '').strip()
        access_days = int(access_days_raw) if access_days_raw.isdigit() else None
        lesson.is_open = True
        lesson.access_days = access_days
        lesson.opened_at = datetime.now(timezone.utc)
        db.session.commit()
        flash('Урок открыт', 'success')
    elif action == 'close':
        lesson.is_open = False
        db.session.commit()
        flash('Урок закрыт', 'success')

    return redirect('/teacher/lessons')


@bp.route('/lessons/<int:lesson_id>/delete', methods=['POST'])
@login_required_teacher
def delete_lesson(lesson_id):
    lesson = db.session.get(Lesson, lesson_id)
    if lesson is None:
        abort(404)
    title = lesson.title
    db.session.delete(lesson)
    db.session.commit()
    flash(f'Урок «{title}» удалён', 'success')
    return redirect('/teacher/lessons')


# ── Task management ────────────────────────────────────────────────────


@bp.route('/lessons/<int:lesson_id>/tasks')
@login_required_teacher
def tasks(lesson_id):
    lesson = db.session.get(Lesson, lesson_id)
    if lesson is None:
        abort(404)
    tasks_list = Task.query.filter_by(lesson_id=lesson_id) \
        .order_by(Task.letter_index).all()
    for t in tasks_list:
        t.tests_count = get_test_count(t.id)
    return render_template('teacher/tasks.html', lesson=lesson, tasks=tasks_list)


@bp.route('/lessons/<int:lesson_id>/tasks/create', methods=['GET', 'POST'])
@login_required_teacher
def create_task(lesson_id):
    lesson = db.session.get(Lesson, lesson_id)
    if lesson is None:
        abort(404)

    if request.method == 'POST':
        short_title = request.form.get('short_title', '').strip()
        problem_text = request.form.get('problem_text', '').strip()
        input_description = request.form.get('input_description', '').strip()
        output_description = request.form.get('output_description', '').strip()
        notes = request.form.get('notes', '').strip()
        time_limit_raw = request.form.get('time_limit', '').strip()
        memory_limit_raw = request.form.get('memory_limit', '').strip()

        existing = Task.query.filter_by(lesson_id=lesson_id).all()

        if len(existing) >= 20:
            flash('Максимум 20 задач на урок', 'error')
            return render_template('teacher/task_form.html', lesson=lesson)

        if not short_title:
            flash('Краткое название обязательно', 'error')
            return render_template('teacher/task_form.html', lesson=lesson)
        if not problem_text:
            flash('Текст условия обязателен', 'error')
            return render_template('teacher/task_form.html', lesson=lesson)
        if not input_description:
            flash('Описание входных данных обязательно', 'error')
            return render_template('teacher/task_form.html', lesson=lesson)
        if not output_description:
            flash('Описание результата обязательно', 'error')
            return render_template('teacher/task_form.html', lesson=lesson)

        existing_letters = [t.letter_index for t in existing]
        next_letter = chr(ord(max(existing_letters)) + 1) if existing_letters else 'A'

        time_limit = int(time_limit_raw) if time_limit_raw.isdigit() else 1
        memory_limit = int(memory_limit_raw) if memory_limit_raw.isdigit() else 256

        t = Task(
            lesson_id=lesson_id,
            letter_index=next_letter,
            short_title=short_title,
            problem_text=problem_text,
            input_description=input_description,
            output_description=output_description,
            notes=notes or None,
            time_limit=time_limit,
            memory_limit=memory_limit,
        )
        db.session.add(t)
        db.session.commit()
        flash(f'Задача {next_letter} создана', 'success')
        return redirect(f'/teacher/lessons/{lesson_id}/tasks')

    return render_template('teacher/task_form.html', lesson=lesson)


@bp.route('/tasks/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required_teacher
def edit_task(task_id):
    task = db.session.get(Task, task_id)
    if task is None:
        abort(404)

    if request.method == 'POST':
        short_title = request.form.get('short_title', '').strip()
        problem_text = request.form.get('problem_text', '').strip()
        input_description = request.form.get('input_description', '').strip()
        output_description = request.form.get('output_description', '').strip()
        notes = request.form.get('notes', '').strip()
        time_limit_raw = request.form.get('time_limit', '').strip()
        memory_limit_raw = request.form.get('memory_limit', '').strip()

        if not short_title:
            flash('Краткое название обязательно', 'error')
            return render_template('teacher/task_form.html', task=task)
        if not problem_text:
            flash('Текст условия обязателен', 'error')
            return render_template('teacher/task_form.html', task=task)
        if not input_description:
            flash('Описание входных данных обязательно', 'error')
            return render_template('teacher/task_form.html', task=task)
        if not output_description:
            flash('Описание результата обязательно', 'error')
            return render_template('teacher/task_form.html', task=task)

        task.short_title = short_title
        task.problem_text = problem_text
        task.input_description = input_description
        task.output_description = output_description
        task.notes = notes or None
        task.time_limit = int(time_limit_raw) if time_limit_raw.isdigit() else 1
        task.memory_limit = int(memory_limit_raw) if memory_limit_raw.isdigit() else 256
        db.session.commit()
        flash('Задача обновлена', 'success')
        return redirect(f'/teacher/lessons/{task.lesson_id}/tasks')

    return render_template('teacher/task_form.html', task=task)


@bp.route('/tasks/<int:task_id>/delete', methods=['POST'])
@login_required_teacher
def delete_task(task_id):
    task = db.session.get(Task, task_id)
    if task is None:
        abort(404)
    lesson_id = task.lesson_id
    letter = task.letter_index
    db.session.delete(task)
    db.session.commit()
    flash(f'Задача {letter} удалена', 'success')
    return redirect(f'/teacher/lessons/{lesson_id}/tasks')


# ── Test file management ───────────────────────────────────────────────


@bp.route('/tasks/<int:task_id>/tests', methods=['GET', 'POST'])
@login_required_teacher
def manage_tests(task_id):
    task = db.session.get(Task, task_id)
    if task is None:
        abort(404)

    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'add':
            input_data = request.form.get('input_data', '')
            expected_output = request.form.get('expected_output', '')
            existing = get_tests(task_id)
            next_num = max((t[0] for t in existing), default=0) + 1
            add_test(task_id, next_num, input_data, expected_output)
            flash(f'Тест {next_num} добавлен', 'success')

        elif action == 'edit':
            test_num = int(request.form.get('test_num', '0'))
            input_data = request.form.get('input_data', '')
            expected_output = request.form.get('expected_output', '')
            update_test(task_id, test_num, input_data, expected_output)
            flash(f'Тест {test_num} обновлён', 'success')

        elif action == 'delete':
            count = get_test_count(task_id)
            if count <= 2:
                flash('Нельзя удалить тест: минимум 2 теста обязательно', 'error')
            else:
                test_num = int(request.form.get('test_num', '0'))
                delete_test(task_id, test_num)
                flash(f'Тест {test_num} удалён', 'success')

        return redirect(f'/teacher/tasks/{task_id}/tests')

    tests_list = get_tests(task_id)
    return render_template('teacher/tests.html', task=task, tests=tests_list)


@bp.route('/students/<int:student_id>/password', methods=['POST'])
@login_required_teacher
def change_student_password(student_id):
    student = db.session.get(Student, student_id)
    if student is None:
        abort(404)

    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        flash('Пароль не может быть пустым', 'error')
        return redirect(f'/teacher/groups/{student.group_id}')

    student.password_plain = new_password
    student.password_hash = generate_password_hash(new_password)
    db.session.commit()
    flash('Пароль изменён', 'success')
    return redirect(f'/teacher/groups/{student.group_id}')


@bp.route('/students/<int:student_id>/reset-password', methods=['POST'])
@login_required_teacher
def reset_student_password(student_id):
    student = db.session.get(Student, student_id)
    if student is None:
        abort(404)

    plain = generate_password()
    student.password_plain = plain
    student.password_hash = generate_password_hash(plain)
    db.session.commit()
    flash(f'Новый пароль: {plain}', 'success')
    return redirect(f'/teacher/groups/{student.group_id}')


@bp.route('/settings', methods=['GET', 'POST'])
@login_required_teacher
def settings():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        teacher = db.session.get(Teacher, session['teacher_id'])

        if not check_password_hash(teacher.password_hash, old_password):
            flash('Неверный старый пароль', 'error')
            return render_template('teacher/settings.html')

        if new_password != confirm_password:
            flash('Пароли не совпадают', 'error')
            return render_template('teacher/settings.html')

        teacher.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('Пароль успешно изменён', 'success')
        return redirect('/teacher/settings')

    return render_template('teacher/settings.html')


# ── Results table and Excel export ─────────────────────────────────────


def _build_results_matrix(group_id):
    group = db.session.get(Group, group_id)
    if group is None:
        return None

    students = Student.query.filter_by(group_id=group_id) \
        .order_by(Student.seq_number).all()

    lessons = Lesson.query.order_by(Lesson.order_number).all()
    tasks = []
    for lesson in lessons:
        lesson_tasks = Task.query.filter_by(lesson_id=lesson.id) \
            .order_by(Task.letter_index).all()
        for t in lesson_tasks:
            tasks.append(t)

    submissions = {}
    for s in students:
        subs = Submission.query.filter_by(student_id=s.id).all()
        submissions[s.id] = {sub.task_id: sub for sub in subs}

    test_counts = {t.id: get_test_count(t.id) for t in tasks}

    matrix = []
    for s in students:
        row = {'student': s, 'cells': []}
        for t in tasks:
            sub = submissions[s.id].get(t.id)
            if sub is None:
                row['cells'].append('')
            else:
                try:
                    result = json.loads(sub.test_results)
                    passed = result.get('passed', 0)
                    total = result.get('total_tests', test_counts[t.id])
                    if result.get('status') == 'ok':
                        row['cells'].append('OK')
                    else:
                        row['cells'].append(f'{passed}/{total}')
                except (json.JSONDecodeError, TypeError):
                    row['cells'].append('')
        matrix.append(row)

    return {
        'group': group,
        'students': students,
        'tasks': tasks,
        'matrix': matrix,
    }


@bp.route('/results')
@login_required_teacher
def results():
    groups = Group.query.order_by(Group.name).all()
    group_id = request.args.get('group', type=int)

    data = None
    if group_id:
        data = _build_results_matrix(group_id)
        if data is None:
            abort(404)

    return render_template('teacher/results.html',
                           groups=groups, group_id=group_id, data=data)


@bp.route('/results/export')
@login_required_teacher
def results_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    group_id = request.args.get('group', type=int)
    if not group_id:
        abort(400)

    data = _build_results_matrix(group_id)
    if data is None:
        abort(404)

    wb = Workbook()
    ws = wb.active
    ws.title = data['group'].name

    bold_font = Font(bold=True)
    center = Alignment(horizontal='center')

    headers = ['Фамилия', 'Имя', 'Логин']
    for t in data['tasks']:
        lesson = db.session.get(Lesson, t.lesson_id)
        header = f'{lesson.title} — {t.letter_index}. {t.short_title}'
        headers.append(header)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.alignment = center

    for row_idx, row in enumerate(data['matrix'], 2):
        s = row['student']
        ws.cell(row=row_idx, column=1, value=s.last_name)
        ws.cell(row=row_idx, column=2, value=s.first_name)
        ws.cell(row=row_idx, column=3, value=s.login)
        for col_idx, cell_val in enumerate(row['cells'], 4):
            ws.cell(row=row_idx, column=col_idx, value=cell_val)

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, len(data['matrix']) + 2):
            val = ws.cell(row=row_idx, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'{data["group"].name}_results.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
